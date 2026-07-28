#!/usr/bin/env python3
"""Deterministically import the Issue #6 source baseline.

The script copies verified source archives, extracts searchable DOCX/CSV/JSON
content, builds the G01-G13 indexes, and normalizes the 71-character / 488-asset
catalogs. It does not create runtime assets or gameplay content.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import io
import json
import re
import shutil
import sys
import zipfile
from dataclasses import dataclass
from pathlib import Path, PurePosixPath
from typing import Any, Iterable
from xml.etree import ElementTree as ET

import openpyxl


W_NS = "http://schemas.openxmlformats.org/wordprocessingml/2006/main"
W = f"{{{W_NS}}}"
REPO_ROOT: Path | None = None


@dataclass(frozen=True)
class PackageSpec:
    package_id: str
    filename: str
    version: str
    purpose: str
    expected_sha256: str
    expected_bytes: int
    authority: str
    source_location: str = "downloads"


PACKAGES: tuple[PackageSpec, ...] = (
    PackageSpec(
        "PKG-PRODUCT-PLAN-V1.1",
        "星骸拾荒者_开发前资料与计划确认包_V1.1_星宇确认版.zip",
        "V1.1",
        "产品、教育、计划与开发前门禁冻结文档",
        "ed51f0f9e6e68eae09fd97fe85b7481b688b93ed06099e0bb1cf9e46090115f2",
        291026,
        "Issue #6 confirmed baseline",
    ),
    PackageSpec(
        "PKG-G02-SCRIPT-FREEZE-V1.0",
        "星骸拾荒者_G02制作脚本总封版与S2启动确认包_V1.0.zip",
        "V1.0",
        "G-S1计划、总封版索引与G-SCR-01—05冻结文本",
        "2cc04b0a73275af35b8b905ee836123af704d04aea1695e943c38d94d535085e",
        463087,
        "Issue #6 confirmed baseline",
    ),
    PackageSpec(
        "PKG-CHARACTERS-V2.1",
        "星骸拾荒者_人物形象设计全集_V2.1_补齐版.zip",
        "V2.1",
        "71人物设计/生产母版与全项目美术资产主清单",
        "a31f21fbe0348be6ff1b9f7b21f53715ccf9dccf59d487cda2296fa4fdd0fceb",
        74253878,
        "SOURCE_PACKAGE_MANIFEST + Issue #6",
    ),
    PackageSpec(
        "PKG-SCENES-V1.0",
        "星骸拾荒者_场景美术设计全集_V1.0.zip",
        "V1.0",
        "91场景设计/生产母版",
        "731cc680ee98eba1bf27474d4d613476dbbe02ca7fff7dfb1beb4b2bb0595de0",
        47905077,
        "SOURCE_PACKAGE_MANIFEST + Issue #6",
        "scene_bundle",
    ),
    PackageSpec(
        "PKG-PROPS-V3.0",
        "星骸拾荒者_道具美术正式包_V3.0.zip",
        "V3.0",
        "PROP-001—046道具设计/生产母版",
        "f712245f25945b234fd73d794b3ff6d6be39744da3a56324a34706ee85f6aa2d",
        31998862,
        "SOURCE_PACKAGE_MANIFEST + Issue #4/#6",
    ),
    PackageSpec(
        "PKG-MECH-V2.0",
        "星骸拾荒者_机制可视化HOPA正式包_V2.0.zip",
        "V2.0",
        "MECH-001—047机制可视化设计/生产母版",
        "de7367d1ec06f97d3b8cca3c671ca9680f522f714929997d4a60fd9af1678b2f",
        39204379,
        "SOURCE_PACKAGE_MANIFEST + Issue #4/#6",
    ),
    PackageSpec(
        "PKG-UI-V2.0",
        "星骸拾荒者_界面美术HOPA正式包_V2.0.zip",
        "V2.0",
        "UI-001—083界面设计/生产母版",
        "827d4262dcf68acf50abdbe77a63192c78cc7b44bf588b389bcbdc612e1d537a",
        49547246,
        "SOURCE_PACKAGE_MANIFEST + Issue #6",
    ),
    PackageSpec(
        "PKG-G02-G13-HOPA-V2.0",
        "星骸拾荒者_G02-G13全章节HOPA实施脚本_V2.0.zip",
        "V2.0",
        "G02—G13完整HOPA实施脚本",
        "f8b9d6f628cac99e5ba799d9cac2f630c45650cf0489dad548de31f569e7eb35",
        762114,
        "SOURCE_PACKAGE_MANIFEST + Issue #4/#6",
    ),
    PackageSpec(
        "PKG-G02-DATA-V2.1",
        "星骸拾荒者_G02锈环星_Unity数据级制作脚本_V2.1.zip",
        "V2.1",
        "G02结构化制作母本（非Unity工程路线）",
        "20d755a73ac1715960cf2c5c5a95b89414d86c9c26c47be76b0a4e6fe5eeb92a",
        119980,
        "Issue #4 supplemental SHA",
    ),
    PackageSpec(
        "PKG-G03-DATA-V2.1",
        "星骸拾荒者_G03齿轮荒原_Unity数据级制作脚本_V2.1.zip",
        "V2.1",
        "G03结构化制作母本（非Unity工程路线）",
        "409ab1be9a180d60eec1d4a15cb6917f628ae4d06d76ee34752d0f621185035b",
        114204,
        "locally observed supplemental package",
    ),
    PackageSpec(
        "PKG-HOPA-FX001-V1.0",
        "星骸拾荒者_HOPA架构与FX-001确认包_V1.0.zip",
        "V1.0",
        "HOPA架构、FX-001全文与四张设计/生产母版",
        "87e1c9f66c5c674c598f62ed69488d6417698aae78666cf749751cb8e93c4ae8",
        1282627,
        "Issue #4 supplemental SHA + Issue #6",
    ),
    PackageSpec(
        "PKG-G01-V3.0",
        "星骸拾荒者_G01序章全量补齐与G01-G13整合正式包_V3.0.zip",
        "V3.0",
        "G01完整剧情、结构化数据、33项正式资产与G02开场边界V2.2",
        "85a20020d471e6dc77454b90e9d7792216db2555aef4e44fa729862ae9ddc043",
        2901474,
        "P0-A_MISSING_SOURCE_HANDOFF",
    ),
    PackageSpec(
        "PKG-G02-G13-DATA-V2.1",
        "星骸拾荒者_G02-G13_Unity数据级制作脚本完整包_V2.1.zip",
        "V2.1",
        "G02—G13完整结构化数据母本；运行时仍为HTML5/PWA",
        "4db2cbb67e688aa7b55b5fe509f38377577e25a2259df3011c105f7c52c59708",
        1112954,
        "P0-A_MISSING_SOURCE_HANDOFF",
    ),
    PackageSpec(
        "PKG-FX-V2.0",
        "星骸拾荒者_技能与装备效果HOPA正式包_V2.0.zip",
        "V2.0",
        "FX-001—041正式清单、设计板与详细规范",
        "882e933ca90917c37a6cd3c88d5988a2ce0ce8c6dbaf1d86e5f29102290ee5e1",
        30540605,
        "P0-A_MISSING_SOURCE_HANDOFF",
    ),
    PackageSpec(
        "PKG-DANGER-V2.0",
        "星骸拾荒者_危险视觉HOPA正式包_V2.0.zip",
        "V2.0",
        "DANGER-001—076正式清单、设计板与详细规范",
        "981d9069efbc7627d4d64dbabd5795acb85f2a4098d9061ba7717d924669181b",
        61745430,
        "P0-A_MISSING_SOURCE_HANDOFF",
    ),
    PackageSpec(
        "DOC-G-S2-D01-V1.0",
        "G-S2-D01_S2视觉总方向关键决策冻结记录_V1.0.docx",
        "V1.0",
        "S2视觉总方向关键决策冻结记录",
        "b746bec0e860b5a457b701d071d33f01316a1c6cc8f3205b06e92e3656b79a77",
        40261,
        "P0-A_MISSING_SOURCE_HANDOFF",
    ),
    PackageSpec(
        "DOC-G-S2-CHG-01-V1.0",
        "G-S2-CHG-01_视觉路线与角色一致性修正记录_V1.0.docx",
        "V1.0",
        "视觉路线与角色一致性修正记录",
        "a62e1c7df26ed537265dda4830cd7f3690d85e79046630869644e8b6e3641cb0",
        39919,
        "P0-A_MISSING_SOURCE_HANDOFF",
    ),
    PackageSpec(
        "DOC-G-CHAR-01-V1.0",
        "G-CHAR-01_主角星宇造型基线_V1.0_C方案布偶修正版.docx",
        "V1.0",
        "主角星宇造型冻结基线",
        "9d5a83301f9462ffd4d268048e4ed809680309b2cb78f9f244d6d3facdddd435",
        40203,
        "P0-A_MISSING_SOURCE_HANDOFF",
    ),
    PackageSpec(
        "DOC-G-ANIM-01-V1.0",
        "G-ANIM-01_图片解密游戏轻动效与骨骼动画方案_V1.0.docx",
        "V1.0",
        "图片解密游戏轻动效与骨骼动画冻结方案",
        "c44dcfef7f8b15176f420eb267c50c0ad6aa0b8819f4ed9e26e8f86c333f9e28",
        40300,
        "P0-A_MISSING_SOURCE_HANDOFF",
    ),
)


_RESOLVED_P0A_GAPS: tuple[dict[str, Any], ...] = (
    {
        "source_id": "PKG-G01-V3.0",
        "name": "G01整合正式包 V3.0",
        "expected_filename": "星骸拾荒者_G01序章全量补齐与G01-G13整合正式包_V3.0.zip",
        "expected_sha256": "85a20020d471e6dc77454b90e9d7792216db2555aef4e44fa729862ae9ddc043",
        "expected_bytes": 2901474,
        "impact": "G01全文、结构化数据与33项新增资产的正式名称/ID无法导入",
    },
    {
        "source_id": "PKG-G02-G13-DATA-V2.1",
        "name": "G02—G13数据级制作脚本完整包 V2.1",
        "expected_filename": "星骸拾荒者_G02-G13_Unity数据级制作脚本完整包_V2.1.zip",
        "expected_sha256": "4db2cbb67e688aa7b55b5fe509f38377577e25a2259df3011c105f7c52c59708",
        "expected_bytes": 1112954,
        "impact": "仅G02和G03独立数据包可用；G04—G13结构化母本缺失",
    },
    {
        "source_id": "PKG-FX-V2.0",
        "name": "技能/装备效果HOPA正式包 V2.0",
        "expected_filename": "星骸拾荒者_技能与装备效果HOPA正式包_V2.0.zip",
        "expected_sha256": "882e933ca90917c37a6cd3c88d5988a2ce0ce8c6dbaf1d86e5f29102290ee5e1",
        "expected_bytes": 30540605,
        "impact": "FX-001—041只能由美术主清单建立索引，无法保留V2设计板原包",
    },
    {
        "source_id": "PKG-DANGER-V2.0",
        "name": "危险视觉HOPA正式包 V2.0",
        "expected_filename": "星骸拾荒者_危险视觉HOPA正式包_V2.0.zip",
        "expected_sha256": "981d9069efbc7627d4d64dbabd5795acb85f2a4098d9061ba7717d924669181b",
        "expected_bytes": 61745430,
        "impact": "DANGER-001—076只能建立可追踪初始槽位，无法保留V2设计板原包",
    },
    {
        "source_id": "DOC-G-S2-D01-V1.0",
        "name": "G-S2-D01 S2视觉总方向关键决策清单 V1.0",
        "expected_filename": "G-S2-D01_*_V1.0*.docx",
        "expected_sha256": None,
        "expected_bytes": None,
        "impact": "仅发现V0.9待确认版，不能冒充冻结V1.0",
    },
    {
        "source_id": "DOC-G-S2-CHG-01-V1.0",
        "name": "G-S2-CHG-01 V1.0",
        "expected_filename": "G-S2-CHG-01_*_V1.0*.docx",
        "expected_sha256": None,
        "expected_bytes": None,
        "impact": "正式视觉变更记录缺失",
    },
    {
        "source_id": "DOC-G-CHAR-01-V1.0",
        "name": "G-CHAR-01 星宇布偶C方案 V1.0",
        "expected_filename": "G-CHAR-01_*_V1.0*.docx",
        "expected_sha256": None,
        "expected_bytes": None,
        "impact": "星宇布偶C方案仅登记为已确认，正式文本未找到",
    },
    {
        "source_id": "DOC-G-ANIM-01-V1.0",
        "name": "G-ANIM-01 V1.0",
        "expected_filename": "G-ANIM-01_*_V1.0*.docx",
        "expected_sha256": None,
        "expected_bytes": None,
        "impact": "正式动画规范文本未找到",
    },
    {
        "source_id": "DOC-G02-BOUNDARY-V2.2",
        "name": "G02开场边界 V2.2",
        "expected_filename": "*G02*边界*V2.2*.docx",
        "expected_sha256": None,
        "expected_bytes": None,
        "impact": "无法导入独立V2.2原文；仓库现有06_G01_G02_BOUNDARY仅作为确认规则摘要",
    },
)

# P0-A is a merge gate: once this importer succeeds, no required source may
# remain missing. The historical gap records above are retained only as an
# audit trail for the remediation diff.
MISSING_REQUIRED: tuple[dict[str, Any], ...] = ()


PRODUCT_DOCS = {
    "G-PROD-01_": "G-PROD-01_V1.1.md",
    "G-EDU-01_": "G-EDU-01_V1.0.md",
    "G-GDD-G02_": "G-GDD-G02_V1.1.md",
    "G-PLAN-01_": "G-PLAN-01_V1.1.md",
    "G-DOC-GATE-01_": "G-DOC-GATE-01_V1.3.md",
    "G-ARCH-M01_": "G-ARCH-M01_V4.3.md",
    "G-DOC-CONF-01_": "G-DOC-CONF-01_V1.0.md",
}

G02_FREEZE_DOCS = {
    "G-S1-PLAN-01_": "G-S1-PLAN-01_V2.0.md",
    "G-S1-CLOSE-01_": "G-S1-CLOSE-01_V1.0.md",
    "G-SCR-01_": "G-SCR-01_V1.0.md",
    "G-SCR-02_": "G-SCR-02_V1.0.md",
    "G-SCR-03_": "G-SCR-03_V1.0.md",
    "G-SCR-04_": "G-SCR-04_V1.0.md",
    "G-SCR-05_": "G-SCR-05_V1.0.md",
}


def sha256_bytes(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def sha256_canonical_text_file(path: Path) -> str:
    """Hash text outputs after normalizing line endings to repository LF."""
    data = path.read_bytes().replace(b"\r\n", b"\n").replace(b"\r", b"\n")
    return sha256_bytes(data)


def repository_path(path: Path) -> str:
    if REPO_ROOT is None:
        raise RuntimeError("repository root has not been initialized")
    return path.resolve().relative_to(REPO_ROOT).as_posix()


def write_json(path: Path, value: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        json.dumps(value, ensure_ascii=False, indent=2, sort_keys=False) + "\n",
        encoding="utf-8",
    )


def write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    headers: list[str] = []
    for row in rows:
        for key in row:
            if key not in headers:
                headers.append(key)
    with path.open("w", encoding="utf-8-sig", newline="") as stream:
        writer = csv.DictWriter(stream, fieldnames=headers, lineterminator="\n")
        writer.writeheader()
        writer.writerows(rows)


def package_path(spec: PackageSpec, downloads: Path) -> Path:
    if spec.source_location == "scene_bundle":
        nested = downloads / "星骸拾荒者_美术设计全集" / spec.filename
        if nested.is_file():
            return nested
    return downloads / spec.filename


def verify_and_copy_packages(
    downloads: Path, originals_dir: Path
) -> tuple[dict[str, Path], list[dict[str, Any]]]:
    originals_dir.mkdir(parents=True, exist_ok=True)
    paths: dict[str, Path] = {}
    manifest: list[dict[str, Any]] = []
    errors: list[str] = []
    for spec in PACKAGES:
        source = package_path(spec, downloads)
        if not source.is_file():
            errors.append(f"missing local source: {source}")
            continue
        observed_size = source.stat().st_size
        observed_sha = sha256_file(source)
        if observed_size != spec.expected_bytes:
            errors.append(
                f"{spec.filename}: bytes {observed_size} != {spec.expected_bytes}"
            )
        if observed_sha != spec.expected_sha256:
            errors.append(
                f"{spec.filename}: sha256 {observed_sha} != {spec.expected_sha256}"
            )
        destination = originals_dir / spec.filename
        if not destination.exists() or sha256_file(destination) != observed_sha:
            shutil.copy2(source, destination)
        paths[spec.package_id] = destination
        manifest.append(
            {
                "package_id": spec.package_id,
                "filename": spec.filename,
                "version": spec.version,
                "purpose": spec.purpose,
                "authority": spec.authority,
                "status": "imported_verified",
                "expected_bytes": spec.expected_bytes,
                "observed_bytes": observed_size,
                "expected_sha256": spec.expected_sha256,
                "observed_sha256": observed_sha,
                "repository_path": repository_path(destination),
                "original_discovery_path": (
                    f"Downloads/星骸拾荒者_美术设计全集/{spec.filename}"
                    if spec.source_location == "scene_bundle"
                    else f"Downloads/{spec.filename}"
                ),
                "storage": "git_lfs",
            }
        )
    if errors:
        raise RuntimeError("\n".join(errors))
    return paths, manifest


def style_names_from_docx(archive: zipfile.ZipFile) -> dict[str, str]:
    if "word/styles.xml" not in archive.namelist():
        return {}
    root = ET.fromstring(archive.read("word/styles.xml"))
    styles: dict[str, str] = {}
    for style in root.findall(f".//{W}style"):
        style_id = style.get(f"{W}styleId")
        name = style.find(f"{W}name")
        if style_id and name is not None:
            styles[style_id] = name.get(f"{W}val", style_id)
    return styles


def paragraph_text(paragraph: ET.Element) -> str:
    pieces: list[str] = []
    for element in paragraph.iter():
        if element.tag == f"{W}t":
            pieces.append(element.text or "")
        elif element.tag == f"{W}tab":
            pieces.append("\t")
        elif element.tag in {f"{W}br", f"{W}cr"}:
            pieces.append("\n")
    return "".join(pieces).strip()


def paragraph_markdown(paragraph: ET.Element, styles: dict[str, str]) -> str:
    text = paragraph_text(paragraph)
    if not text:
        return ""
    style = paragraph.find(f"./{W}pPr/{W}pStyle")
    style_name = ""
    if style is not None:
        style_id = style.get(f"{W}val", "")
        style_name = styles.get(style_id, style_id).lower()
    heading_match = re.search(r"(?:heading|标题)\s*([1-6])", style_name)
    if heading_match:
        return f"{'#' * int(heading_match.group(1))} {text}"
    return text


def table_markdown(table: ET.Element, styles: dict[str, str]) -> str:
    rows: list[list[str]] = []
    for tr in table.findall(f"./{W}tr"):
        row: list[str] = []
        for tc in tr.findall(f"./{W}tc"):
            paragraphs = [
                paragraph_markdown(p, styles).lstrip("# ").strip()
                for p in tc.findall(f".//{W}p")
            ]
            cell = "<br>".join(value for value in paragraphs if value)
            row.append(cell.replace("|", "\\|"))
        if row:
            rows.append(row)
    if not rows:
        return ""
    width = max(len(row) for row in rows)
    padded = [row + [""] * (width - len(row)) for row in rows]
    output = [
        "| " + " | ".join(padded[0]) + " |",
        "| " + " | ".join(["---"] * width) + " |",
    ]
    output.extend("| " + " | ".join(row) + " |" for row in padded[1:])
    return "\n".join(output)


def document_blocks(container: ET.Element, styles: dict[str, str]) -> Iterable[str]:
    for child in list(container):
        if child.tag == f"{W}p":
            value = paragraph_markdown(child, styles)
            if value:
                yield value
        elif child.tag == f"{W}tbl":
            value = table_markdown(child, styles)
            if value:
                yield value
        else:
            yield from document_blocks(child, styles)


def docx_to_markdown(
    data: bytes,
    *,
    package_id: str,
    source_entry: str,
    version: str,
    purpose: str,
) -> tuple[str, dict[str, Any]]:
    with zipfile.ZipFile(io.BytesIO(data)) as archive:
        root = ET.fromstring(archive.read("word/document.xml"))
        styles = style_names_from_docx(archive)
    body = root.find(f".//{W}body")
    if body is None:
        raise ValueError(f"DOCX has no word body: {source_entry}")
    blocks = list(document_blocks(body, styles))
    xml_text_nodes = root.findall(f".//{W}t")
    source_chars = sum(len(node.text or "") for node in xml_text_nodes)
    captured_chars = sum(
        len(node.text or "") for node in body.findall(f".//{W}t")
    )
    table_count = len(root.findall(f".//{W}tbl"))
    source_hash = sha256_bytes(data)
    metadata = (
        "---\n"
        f"source_package: {package_id}\n"
        f"source_entry: {json.dumps(source_entry, ensure_ascii=False)}\n"
        f"source_sha256: {source_hash}\n"
        f"source_bytes: {len(data)}\n"
        f"table_count: {table_count}\n"
        f"version: {version}\n"
        f"purpose: {json.dumps(purpose, ensure_ascii=False)}\n"
        "extraction: full_text_and_tables_from_ooxml\n"
        "runtime_asset: false\n"
        "---\n\n"
    )
    markdown = metadata + "\n\n".join(blocks).strip() + "\n"
    stats = {
        "source_package": package_id,
        "original_filename": PurePosixPath(source_entry).name,
        "source_entry": source_entry,
        "source_sha256": source_hash,
        "source_bytes": len(data),
        "source_wt_nodes": len(xml_text_nodes),
        "source_text_characters": source_chars,
        "captured_text_characters": captured_chars,
        "table_count": table_count,
        "coverage_ratio": 1 if source_chars == captured_chars else 0,
        "markdown_characters": len(markdown),
        "extraction": "full_text_and_tables_from_ooxml",
    }
    return markdown, stats


def find_entry(archive: zipfile.ZipFile, predicate: Any) -> str:
    matches = [name for name in archive.namelist() if predicate(name)]
    if len(matches) != 1:
        raise ValueError(f"expected one ZIP entry, found {len(matches)}: {matches}")
    return matches[0]


def record_extraction(
    records: list[dict[str, Any]],
    *,
    output: Path,
    package_id: str,
    entry: str,
    source_data: bytes,
    extraction: str,
) -> None:
    output_hash_mode = (
        "raw_bytes"
        if extraction in {"verbatim_design_production_master", "verbatim_binary"}
        else "canonical_lf"
    )
    output_sha256 = (
        sha256_file(output)
        if output_hash_mode == "raw_bytes"
        else sha256_canonical_text_file(output)
    )
    records.append(
        {
            "output_path": repository_path(output),
            "output_sha256": output_sha256,
            "output_hash_mode": output_hash_mode,
            "source_package": package_id,
            "source_entry": entry,
            "source_entry_sha256": sha256_bytes(source_data),
            "extraction": extraction,
        }
    )


def record_derived_catalog(
    records: list[dict[str, Any]],
    *,
    output: Path,
    derived_from: list[dict[str, Any]],
    field_authority_map: dict[str, Any],
) -> None:
    records.append(
        {
            "output_path": repository_path(output),
            "output_sha256": sha256_canonical_text_file(output),
            "output_hash_mode": "canonical_lf",
            "extraction": "multi_source_derived_catalog",
            "generated_by": CATALOG_GENERATOR,
            "mapping_version": CATALOG_MAPPING_VERSION,
            "registry_reference_role": REGISTRY_REFERENCE_ROLE,
            "field_authority_map": field_authority_map,
            "derived_from": derived_from,
        }
    )


def extract_docx_entry(
    package: Path,
    package_id: str,
    entry: str,
    output: Path,
    version: str,
    purpose: str,
    records: list[dict[str, Any]],
) -> dict[str, Any]:
    with zipfile.ZipFile(package) as archive:
        data = archive.read(entry)
    markdown, stats = docx_to_markdown(
        data,
        package_id=package_id,
        source_entry=entry,
        version=version,
        purpose=purpose,
    )
    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text(markdown, encoding="utf-8")
    record_extraction(
        records,
        output=output,
        package_id=package_id,
        entry=entry,
        source_data=data,
        extraction="docx_full_text_and_tables_to_markdown",
    )
    stats["output_path"] = repository_path(output)
    return stats


def extract_standalone_docx(
    source: Path,
    source_id: str,
    output: Path,
    purpose: str,
    records: list[dict[str, Any]],
) -> dict[str, Any]:
    data = source.read_bytes()
    markdown, stats = docx_to_markdown(
        data,
        package_id=source_id,
        source_entry=source.name,
        version="V1.0",
        purpose=purpose,
    )
    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text(markdown, encoding="utf-8")
    record_extraction(
        records,
        output=output,
        package_id=source_id,
        entry=source.name,
        source_data=data,
        extraction="full_text_and_tables_from_ooxml",
    )
    stats["output_path"] = repository_path(output)
    stats["source_repository_path"] = repository_path(source)
    return stats


def extract_named_docs(
    package: Path,
    package_id: str,
    mapping: dict[str, str],
    output_dir: Path,
    version: str,
    purpose: str,
    records: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    stats: list[dict[str, Any]] = []
    with zipfile.ZipFile(package) as archive:
        names = archive.namelist()
    for prefix, output_name in mapping.items():
        entry = next(
            (
                name
                for name in names
                if PurePosixPath(name).name.startswith(prefix)
                and name.lower().endswith(".docx")
            ),
            None,
        )
        if entry is None:
            raise ValueError(f"{package.name}: missing required entry prefix {prefix}")
        stats.append(
            extract_docx_entry(
                package,
                package_id,
                entry,
                output_dir / output_name,
                version,
                purpose,
                records,
            )
        )
    return stats


def extract_story_docs(
    package: Path,
    g01_package: Path,
    repo: Path,
    records: list[dict[str, Any]],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    story_dir = repo / "docs/story/G01-G13"
    story_dir.mkdir(parents=True, exist_ok=True)
    with zipfile.ZipFile(g01_package) as archive:
        g01_entry = find_entry(
            archive,
            lambda name: name.lower().endswith(".docx")
            and "G01拾光号坠落之前_剧情与HOPA实施脚本_V3.0" in name,
        )
    g01_output = story_dir / "G01.md"
    g01_stats = extract_docx_entry(
        g01_package,
        "PKG-G01-V3.0",
        g01_entry,
        g01_output,
        "V3.0",
        "G01八场景完整剧情与HOPA实施脚本正式全文",
        records,
    )
    stats: list[dict[str, Any]] = [{"chapter": "G01", **g01_stats}]
    index: list[dict[str, Any]] = [
        {
            "chapter": "G01",
            "status": "imported_verified",
            "version": "V3.0",
            "path": "docs/story/G01-G13/G01.md",
            "source_entry": g01_entry,
            "source_sha256": g01_stats["source_sha256"],
            "coverage_ratio": g01_stats["coverage_ratio"],
            "g01_chapter_complete": True,
            "g01_handoff_to_g02": True,
            "world_star_core_count": 0,
        }
    ]
    with zipfile.ZipFile(package) as archive:
        entries = [
            name
            for name in archive.namelist()
            if name.lower().endswith(".docx")
        ]
    for chapter_number in range(2, 14):
        chapter = f"G{chapter_number:02d}"
        entry = next(
            (
                name
                for name in entries
                if re.search(fr"_{chapter}_", PurePosixPath(name).name)
                and "更新总说明" not in name
            ),
            None,
        )
        if entry is None:
            raise ValueError(f"missing story DOCX for {chapter}")
        output = story_dir / f"{chapter}.md"
        chapter_stats = extract_docx_entry(
            package,
            "PKG-G02-G13-HOPA-V2.0",
            entry,
            output,
            "V2.0",
            f"{chapter} HOPA实施脚本正式全文",
            records,
        )
        stats.append({"chapter": chapter, **chapter_stats})
        index.append(
            {
                "chapter": chapter,
                "status": "imported_verified",
                "version": "V2.0",
                "path": output.relative_to(repo).as_posix(),
                "source_entry": entry,
                "source_sha256": chapter_stats["source_sha256"],
                "coverage_ratio": chapter_stats["coverage_ratio"],
            }
        )
    with zipfile.ZipFile(package) as archive:
        notes_entry = next(
            name for name in entries if "更新总说明" in name
        )
    stats.append(
        {
            "chapter": "G02-G13-NOTES",
            **extract_docx_entry(
                package,
                "PKG-G02-G13-HOPA-V2.0",
                notes_entry,
                story_dir / "HOPA_UPDATE_NOTES_V2.0.md",
                "V2.0",
                "G02—G13全章节脚本更新总说明",
                records,
            ),
        }
    )
    write_json(story_dir / "index.json", index)
    return stats, index


def extract_data_package(
    package: Path,
    package_id: str,
    chapter: str,
    repo: Path,
    records: list[dict[str, Any]],
) -> dict[str, Any]:
    output_root = repo / "data/source/g02-g13" / chapter
    output_root.mkdir(parents=True, exist_ok=True)
    with zipfile.ZipFile(package) as archive:
        names = archive.namelist()
        docx_entry = find_entry(
            archive,
            lambda name: name.lower().endswith(".docx") and "制作脚本说明" in name,
        )
    doc_stats = extract_docx_entry(
        package,
        package_id,
        docx_entry,
        output_root / "README.md",
        "V2.1",
        f"{chapter}结构化制作母本说明；不代表Unity运行时路线",
        records,
    )
    file_index: list[dict[str, Any]] = []
    with zipfile.ZipFile(package) as archive:
        for entry in names:
            normalized = PurePosixPath(entry)
            if "02_CSV数据表" in normalized.parts and entry.lower().endswith(".csv"):
                destination = output_root / "csv" / normalized.name
                extraction = "verbatim_csv"
            elif "03_JSON数据" in normalized.parts and entry.lower().endswith(".json"):
                destination = output_root / "json" / normalized.name
                extraction = "verbatim_json"
            else:
                continue
            data = archive.read(entry)
            destination.parent.mkdir(parents=True, exist_ok=True)
            destination.write_bytes(data)
            record_extraction(
                records,
                output=destination,
                package_id=package_id,
                entry=entry,
                source_data=data,
                extraction=extraction,
            )
            file_index.append(
                {
                    "path": destination.relative_to(repo).as_posix(),
                    "source_entry": entry,
                    "sha256": sha256_bytes(data),
                    "format": destination.suffix.lstrip("."),
                }
            )
    write_json(output_root / "index.json", file_index)
    return {
        "chapter": chapter,
        "status": "imported_verified",
        "version": "V2.1",
        "package_id": package_id,
        "description_coverage_ratio": doc_stats["coverage_ratio"],
        "files": len(file_index),
        "path": output_root.relative_to(repo).as_posix(),
    }


def extract_complete_chapter_data(
    package: Path,
    package_id: str,
    chapter: str,
    repo: Path,
    records: list[dict[str, Any]],
    *,
    version: str,
    g01: bool = False,
    structured_package: Path | None = None,
    structured_package_id: str | None = None,
) -> dict[str, Any]:
    output_root = repo / ("data/source/g01" if g01 else f"data/source/g02-g13/{chapter}")
    output_root.mkdir(parents=True, exist_ok=True)
    with zipfile.ZipFile(package) as archive:
        names = archive.namelist()
        if g01:
            scoped = [name for name in names if "03_G01_Unity数据级脚本" in name]
        else:
            scoped = [
                name
                for name in names
                if any(part.startswith(f"{chapter}_") for part in PurePosixPath(name).parts)
            ]
        docx_entries = [
            name
            for name in scoped
            if name.lower().endswith(".docx")
            and (
                ("Unity实施说明" in name)
                if g01
                else ("制作脚本说明" in name)
            )
        ]
        if len(docx_entries) != 1:
            raise ValueError(f"{chapter}: expected one chapter README DOCX, got {docx_entries}")
        workbook_entries = [
            name
            for name in scoped
            if name.lower().endswith(".xlsx") and "制作脚本" in name
        ]
        if len(workbook_entries) != 1:
            raise ValueError(f"{chapter}: expected one chapter workbook, got {workbook_entries}")

    doc_stats = extract_docx_entry(
        package,
        package_id,
        docx_entries[0],
        output_root / "README.md",
        version,
        f"{chapter}结构化制作数据母本说明；正式运行时为HTML5/PWA",
        records,
    )
    file_index: list[dict[str, Any]] = []
    with zipfile.ZipFile(package) as archive:
        for entry in scoped:
            normalized = PurePosixPath(entry)
            if entry == workbook_entries[0]:
                destination = output_root / "master" / normalized.name
                extraction = "verbatim_binary"
            else:
                continue
            data = archive.read(entry)
            destination.parent.mkdir(parents=True, exist_ok=True)
            destination.write_bytes(data)
            record_extraction(
                records,
                output=destination,
                package_id=package_id,
                entry=entry,
                source_data=data,
                extraction=extraction,
            )
            file_index.append(
                {
                    "path": destination.relative_to(repo).as_posix(),
                    "package_id": package_id,
                    "source_entry": entry,
                    "sha256": sha256_bytes(data),
                    "format": destination.suffix.lstrip("."),
                    "extraction": extraction,
                }
            )

    data_package = structured_package or package
    data_package_id = structured_package_id or package_id
    with zipfile.ZipFile(data_package) as archive:
        data_names = archive.namelist()
        if structured_package is None:
            data_scoped = scoped
        else:
            data_scoped = data_names
        for entry in data_scoped:
            normalized = PurePosixPath(entry)
            if "02_CSV数据表" in normalized.parts and entry.lower().endswith(".csv"):
                destination = output_root / "csv" / normalized.name
                extraction = "verbatim_csv"
            elif "03_JSON数据" in normalized.parts and entry.lower().endswith(".json"):
                destination = output_root / "json" / normalized.name
                extraction = "verbatim_json"
            else:
                continue
            data = archive.read(entry)
            destination.parent.mkdir(parents=True, exist_ok=True)
            destination.write_bytes(data)
            record_extraction(
                records,
                output=destination,
                package_id=data_package_id,
                entry=entry,
                source_data=data,
                extraction=extraction,
            )
            file_index.append(
                {
                    "path": destination.relative_to(repo).as_posix(),
                    "package_id": data_package_id,
                    "source_entry": entry,
                    "sha256": sha256_bytes(data),
                    "format": destination.suffix.lstrip("."),
                    "extraction": extraction,
                }
            )

    required_names = (
        "场景流程",
        "热点清单",
        "找物清单",
        "背包道具流转",
        "对话脚本",
        "场景状态机",
        "三级提示",
        "存档与恢复",
        "程序变量",
        "资产映射",
    )
    missing_structured = [
        f"{name}{suffix}"
        for name in required_names
        for suffix in (".csv", ".json")
        if not any(
            PurePosixPath(item["path"]).name == f"{name}{suffix}"
            for item in file_index
        )
    ]
    master_name = f"{chapter}_MasterData.json"
    if not any(
        PurePosixPath(item["path"]).name == master_name for item in file_index
    ):
        missing_structured.append(master_name)

    if missing_structured:
        with zipfile.ZipFile(package) as archive:
            workbook_data = archive.read(workbook_entries[0])
        workbook = openpyxl.load_workbook(io.BytesIO(workbook_data), data_only=True)
        modules: dict[str, list[dict[str, Any]]] = {}
        for sheet_name in required_names:
            if sheet_name not in workbook.sheetnames:
                raise ValueError(f"{chapter}: workbook missing required sheet {sheet_name}")
            sheet = workbook[sheet_name]
            headers = [
                str(value).strip() if value is not None else ""
                for value in next(
                    sheet.iter_rows(min_row=3, max_row=3, values_only=True)
                )
            ]
            rows: list[dict[str, Any]] = []
            for values in sheet.iter_rows(min_row=4, values_only=True):
                row = {
                    header: values[index]
                    for index, header in enumerate(headers)
                    if header and index < len(values) and values[index] is not None
                }
                if row:
                    rows.append(row)
            if not rows:
                raise ValueError(f"{chapter}: workbook sheet {sheet_name} has no data")
            modules[sheet_name] = rows
            for suffix in (".csv", ".json"):
                destination = output_root / suffix.lstrip(".") / f"{sheet_name}{suffix}"
                if suffix == ".csv":
                    write_csv(destination, rows)
                    extraction = "xlsx_sheet_to_csv"
                else:
                    write_json(destination, rows)
                    extraction = "xlsx_sheet_to_json"
                output_data = destination.read_bytes()
                record_extraction(
                    records,
                    output=destination,
                    package_id=package_id,
                    entry=workbook_entries[0],
                    source_data=workbook_data,
                    extraction=extraction,
                )
                file_index.append(
                    {
                        "path": destination.relative_to(repo).as_posix(),
                        "package_id": package_id,
                        "source_entry": workbook_entries[0],
                        "source_sha256": sha256_bytes(workbook_data),
                        "sha256": sha256_bytes(output_data),
                        "format": suffix.lstrip("."),
                        "extraction": extraction,
                        "source_sheet": sheet_name,
                    }
                )
        chapter_name_match = re.search(
            rf"{chapter}(.+?)_Unity", PurePosixPath(workbook_entries[0]).name
        )
        master_data = {
            "chapter": chapter,
            "chapter_name": (
                chapter_name_match.group(1) if chapter_name_match else chapter
            ),
            "version": version,
            "protagonist": "星宇",
            "source_workbook": workbook_entries[0],
            "source_package_id": package_id,
            "extraction": "deterministic_workbook_transformation",
            "modules": modules,
        }
        master_destination = output_root / "json" / master_name
        write_json(master_destination, master_data)
        output_data = master_destination.read_bytes()
        record_extraction(
            records,
            output=master_destination,
            package_id=package_id,
            entry=workbook_entries[0],
            source_data=workbook_data,
            extraction="xlsx_workbook_to_masterdata",
        )
        file_index.append(
            {
                "path": master_destination.relative_to(repo).as_posix(),
                "package_id": package_id,
                "source_entry": workbook_entries[0],
                "source_sha256": sha256_bytes(workbook_data),
                "sha256": sha256_bytes(output_data),
                "format": "json",
                "extraction": "xlsx_workbook_to_masterdata",
            }
        )

    for name in required_names:
        for suffix in (".csv", ".json"):
            if not any(PurePosixPath(item["path"]).name == f"{name}{suffix}" for item in file_index):
                raise ValueError(f"{chapter}: missing {name}{suffix}")
    if not any(PurePosixPath(item["path"]).name == master_name for item in file_index):
        raise ValueError(f"{chapter}: missing {master_name}")
    write_json(output_root / "index.json", file_index)
    return {
        "chapter": chapter,
        "status": "imported_verified",
        "version": version,
        "package_id": package_id,
        "structured_package_id": data_package_id,
        "structured_data_origin": (
            "verbatim_package_files"
            if not missing_structured
            else "deterministic_formal_workbook_extraction"
        ),
        "description_coverage_ratio": doc_stats["coverage_ratio"],
        "files": len(file_index),
        "path": output_root.relative_to(repo).as_posix(),
        "runtime_technology": "HTML5/PWA + Vite + TypeScript",
        **(
            {
                "g01_chapter_complete": True,
                "g01_handoff_to_g02": True,
                "world_star_core_count": 0,
            }
            if g01
            else {}
        ),
    }


def xlsx_master_from_character_package(package: Path) -> tuple[bytes, str]:
    with zipfile.ZipFile(package) as archive:
        entry = find_entry(
            archive,
            lambda name: name.endswith("星骸拾荒者_全项目美术资产主清单_V1.0.xlsx"),
        )
        return archive.read(entry), entry


def sheet_rows(
    workbook: openpyxl.Workbook, sheet_name: str, id_prefix: str
) -> list[dict[str, Any]]:
    sheet = workbook[sheet_name]
    header_values = [cell.value for cell in sheet[4]]
    headers = [str(value).strip() if value is not None else "" for value in header_values]
    rows: list[dict[str, Any]] = []
    for values in sheet.iter_rows(min_row=5, values_only=True):
        asset_id = values[0]
        if not isinstance(asset_id, str) or not asset_id.startswith(f"{id_prefix}-"):
            continue
        row: dict[str, Any] = {}
        for index, header in enumerate(headers):
            if not header or index >= len(values):
                continue
            value = values[index]
            if value is not None:
                row[header] = value
        rows.append(row)
    return rows


CATALOG_MAPPING_VERSION = "formal-row-authority-v1"
CATALOG_GENERATOR = "scripts/import_source_baseline.py#build_catalogs"
REGISTRY_REFERENCE_ROLE = "cross_check_only"


def normalize_legacy_asset_row(
    source: dict[str, Any],
    domain: str,
    source_sheet: str,
    source_package: str = "PKG-CHARACTERS-V2.1",
    source_entry: str = "",
    source_sha256: str = "",
    source_status: str = "legacy_catalog_normalization",
) -> dict[str, Any]:
    name = (
        source.get("人物名称")
        or source.get("场景名称")
        or source.get("名称")
        or source.get("资产ID")
    )
    return {
        "catalog_id": source["资产ID"],
        "official_id": source["资产ID"],
        "domain": domain,
        "name": name,
        "chapter": source.get("星球编号") or source.get("范围"),
        "scope": source.get("星球/范围") or source.get("星球") or source.get("范围"),
        "type": source.get("类别") or source.get("类型"),
        "maturity": source.get("当前状态"),
        "priority": source.get("优先级"),
        "source_sheet": source_sheet,
        "source_package": source_package,
        "source_entry": source_entry,
        "source_sha256": source_sha256,
        "source_granularity": "formal_catalog_entry",
        "source_status": source_status,
        "design_master": True,
        "production_spec": True,
        "runtime_asset": False,
        "acceptance_asset": False,
    }


def require_formal_fields(
    row: dict[str, Any],
    *,
    package_id: str,
    source_entry: str,
    required: tuple[str, ...],
) -> None:
    asset_id = row.get("资产ID", "<missing-id>")
    missing = [field for field in required if field not in row or row[field] in (None, "")]
    if missing:
        raise ValueError(
            f"{package_id}:{source_entry}:{asset_id}: missing formal fields {missing}"
        )


def formal_source_fields(
    *,
    package_id: str,
    source_entry: str,
    source_sha256: str,
    row_id: str,
) -> dict[str, Any]:
    return {
        "source_package": package_id,
        "source_entry": source_entry,
        "source_sha256": source_sha256,
        "source_granularity": "formal_catalog_entry",
        "formal_row_id": row_id,
        "source_status": "formal_row_authoritative",
        "design_master": True,
        "production_spec": True,
        "runtime_asset": False,
        "acceptance_asset": False,
    }


def normalize_scene_formal_row(
    row: dict[str, Any], package_id: str, source_entry: str, source_sha256: str
) -> dict[str, Any]:
    require_formal_fields(
        row,
        package_id=package_id,
        source_entry=source_entry,
        required=("资产ID", "星球编号", "星球/范围", "场景名称", "状态", "优先级", "交付状态"),
    )
    asset_id = row["资产ID"]
    return {
        "catalog_id": asset_id,
        "official_id": asset_id,
        "domain": "scene",
        "name": row["场景名称"],
        "chapter": row["星球编号"],
        "scope": row["星球/范围"],
        "maturity": row["状态"],
        "priority": row["优先级"],
        "delivery_status": row["交付状态"],
        **formal_source_fields(
            package_id=package_id,
            source_entry=source_entry,
            source_sha256=source_sha256,
            row_id=asset_id,
        ),
    }


def normalize_prop_formal_row(
    row: dict[str, Any], package_id: str, source_entry: str, source_sha256: str
) -> dict[str, Any]:
    require_formal_fields(
        row,
        package_id=package_id,
        source_entry=source_entry,
        required=("资产ID", "星球编号", "星球", "道具名称", "对应正式板", "状态", "核查结论"),
    )
    asset_id = row["资产ID"]
    return {
        "catalog_id": asset_id,
        "official_id": asset_id,
        "domain": "prop",
        "name": row["道具名称"],
        "chapter": row["星球编号"],
        "scope": row["星球"],
        "design_board": row["对应正式板"],
        "maturity": row["状态"],
        "verification_result": row["核查结论"],
        **formal_source_fields(
            package_id=package_id,
            source_entry=source_entry,
            source_sha256=source_sha256,
            row_id=asset_id,
        ),
    }


def normalize_mech_formal_row(
    row: dict[str, Any], package_id: str, source_entry: str, source_sha256: str
) -> dict[str, Any]:
    require_formal_fields(
        row,
        package_id=package_id,
        source_entry=source_entry,
        required=("资产ID", "星球", "机制名称", "状态", "冻结版本"),
    )
    asset_id = row["资产ID"]
    chapter_match = re.match(r"^(G\d{2})\b", row["星球"])
    return {
        "catalog_id": asset_id,
        "official_id": asset_id,
        "domain": "mechanism",
        "name": row["机制名称"],
        "mechanism_name": row["机制名称"],
        "chapter": chapter_match.group(1) if chapter_match else row["星球"],
        "scope": row["星球"],
        "maturity": row["状态"],
        "freeze_version": row["冻结版本"],
        **formal_source_fields(
            package_id=package_id,
            source_entry=source_entry,
            source_sha256=source_sha256,
            row_id=asset_id,
        ),
    }


def normalize_ui_formal_row(
    row: dict[str, Any], package_id: str, source_entry: str, source_sha256: str
) -> dict[str, Any]:
    require_formal_fields(
        row,
        package_id=package_id,
        source_entry=source_entry,
        required=(
            "资产ID",
            "范围",
            "星球",
            "界面名称",
            "类别",
            "分批",
            "独立设计板",
            "确认状态",
            "冻结版本",
        ),
    )
    asset_id = row["资产ID"]
    return {
        "catalog_id": asset_id,
        "official_id": asset_id,
        "domain": "ui",
        "name": row["界面名称"],
        "chapter": row["范围"],
        "scope": row["星球"],
        "type": row["类别"],
        "category": row["类别"],
        "batch": row["分批"],
        "independent_board_status": row["独立设计板"],
        "confirmation_status": row["确认状态"],
        "maturity": row["确认状态"],
        "freeze_version": row["冻结版本"],
        **formal_source_fields(
            package_id=package_id,
            source_entry=source_entry,
            source_sha256=source_sha256,
            row_id=asset_id,
        ),
    }


def build_catalogs_legacy(
    package: Path,
    repo: Path,
    records: list[dict[str, Any]],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    data, entry = xlsx_master_from_character_package(package)
    workbook = openpyxl.load_workbook(io.BytesIO(data), data_only=False)
    tables = {
        "characters": sheet_rows(workbook, "人物清单", "CHAR"),
        "scenes": sheet_rows(workbook, "场景清单", "SCN"),
        "props": sheet_rows(workbook, "装备道具", "PROP"),
        "fx": sheet_rows(workbook, "技能特效", "FX"),
        "mech": sheet_rows(workbook, "玩法机制", "MECH"),
        "ui": sheet_rows(workbook, "UI清单", "UI"),
        "hazards": sheet_rows(workbook, "敌人与危险", "HAZ"),
    }
    expected = {
        "characters": 71,
        "scenes": 91,
        "props": 46,
        "fx": 41,
        "mech": 47,
        "ui": 83,
        "hazards": 32,
    }
    observed = {key: len(value) for key, value in tables.items()}
    if observed != expected:
        raise ValueError(f"master catalog counts mismatch: {observed} != {expected}")

    catalogs_dir = repo / "data/source/catalogs"
    characters = tables["characters"]
    write_json(catalogs_dir / "characters-71.json", characters)
    write_csv(catalogs_dir / "characters-71.csv", characters)

    assets: list[dict[str, Any]] = []
    assets.extend(
        normalize_legacy_asset_row(row, "character", "人物清单")
        for row in tables["characters"]
    )
    assets.extend(
        normalize_legacy_asset_row(row, "scene", "场景清单") for row in tables["scenes"]
    )
    assets.extend(
        normalize_legacy_asset_row(row, "prop", "装备道具") for row in tables["props"]
    )
    assets.extend(
        normalize_legacy_asset_row(row, "fx", "技能特效") for row in tables["fx"]
    )
    assets.extend(
        normalize_legacy_asset_row(row, "mechanism", "玩法机制")
        for row in tables["mech"]
    )
    assets.extend(normalize_legacy_asset_row(row, "ui", "UI清单") for row in tables["ui"])

    danger_rows: list[dict[str, Any]] = []
    for index, row in enumerate(tables["hazards"], start=1):
        danger_rows.append(
            {
                "catalog_id": f"DANGER-{index:03d}",
                "official_id": f"DANGER-{index:03d}",
                "domain": "danger",
                "name": row.get("名称"),
                "chapter": row.get("范围"),
                "scope": row.get("范围"),
                "type": row.get("类型"),
                "maturity": row.get("当前状态"),
                "priority": None,
                "source_sheet": "敌人与危险",
                "source_status": "normalized_from_HAZ_base_design; V2_package_missing",
                "source_id": row.get("资产ID"),
                "runtime_asset": False,
            }
        )
    shared = tables["hazards"][6:10]
    chapter_names = [
        ("G03", "齿轮荒原"),
        ("G04", "镜面沙海"),
        ("G05", "刻度冰原"),
        ("G06", "分片群岛"),
        ("G07", "方格城"),
        ("G08", "钟摆之城"),
        ("G09", "棋阵卫星"),
        ("G10", "概率云港"),
        ("G11", "立体工厂"),
        ("G12", "百工星环"),
        ("G13", "零号地球"),
    ]
    danger_index = 33
    for source in shared:
        for chapter, planet in chapter_names:
            danger_rows.append(
                {
                    "catalog_id": f"DANGER-{danger_index:03d}",
                    "official_id": f"DANGER-{danger_index:03d}",
                    "domain": "danger",
                    "name": f"{source.get('名称')}·{planet}皮肤",
                    "chapter": chapter,
                    "scope": planet,
                    "type": "共享原型星球化视觉应用",
                    "maturity": source.get("当前状态"),
                    "priority": None,
                    "source_sheet": "敌人与危险",
                    "source_status": (
                        "expanded_from_explicit_4x11_skin_rule; V2_package_missing"
                    ),
                    "source_id": source.get("资产ID"),
                    "runtime_asset": False,
                }
            )
            danger_index += 1
    if len(danger_rows) != 76:
        raise ValueError(f"danger visual catalog must have 76 rows, got {len(danger_rows)}")
    assets.extend(danger_rows)

    for index in range(1, 34):
        assets.append(
            {
                "catalog_id": f"G01-SLOT-{index:03d}",
                "official_id": None,
                "domain": "g01_addition",
                "name": None,
                "chapter": "G01",
                "scope": "序章正式包",
                "type": "待G01 V3.0正式包提供",
                "maturity": "source_missing",
                "priority": None,
                "source_sheet": None,
                "source_status": "inventory_slot_only; no official ID invented",
                "runtime_asset": False,
            }
        )

    domain_counts: dict[str, int] = {}
    for row in assets:
        domain_counts[row["domain"]] = domain_counts.get(row["domain"], 0) + 1
    expected_domains = {
        "character": 71,
        "scene": 91,
        "prop": 46,
        "fx": 41,
        "mechanism": 47,
        "danger": 76,
        "ui": 83,
        "g01_addition": 33,
    }
    if len(assets) != 488 or domain_counts != expected_domains:
        raise ValueError(
            f"asset catalog mismatch: total={len(assets)}, domains={domain_counts}"
        )
    write_json(
        catalogs_dir / "asset-catalog-488.json",
        {
            "schema_version": 1,
            "total": len(assets),
            "domain_counts": domain_counts,
            "source_package": "PKG-CHARACTERS-V2.1",
            "source_entry": entry,
            "caveats": [
                "DANGER V2原包缺失；76项由主清单的32基础项和4×11明确皮肤规则建立可追踪初始索引。",
                "G01 V3.0原包缺失；33项仅建立库存槽位，不生成正式ID、名称或美术。",
                "所有美术条目均为设计/生产母版索引，不是运行时资产。",
            ],
            "assets": assets,
        },
    )
    write_csv(catalogs_dir / "asset-catalog-488.csv", assets)
    write_json(
        catalogs_dir / "master-workbook-counts.json",
        {
            "source_package": "PKG-CHARACTERS-V2.1",
            "source_entry": entry,
            "source_entry_sha256": sha256_bytes(data),
            "counts": observed,
            "normalized_asset_total": 488,
            "design_or_production_master": True,
            "runtime_asset": False,
        },
    )
    for output in (
        catalogs_dir / "characters-71.json",
        catalogs_dir / "characters-71.csv",
        catalogs_dir / "asset-catalog-488.json",
        catalogs_dir / "asset-catalog-488.csv",
        catalogs_dir / "master-workbook-counts.json",
    ):
        record_extraction(
            records,
            output=output,
            package_id="PKG-CHARACTERS-V2.1",
            entry=entry,
            source_data=data,
            extraction="xlsx_table_normalization",
        )
    return characters, assets


def build_catalogs(
    character_package: Path,
    scene_package: Path,
    prop_package: Path,
    mech_package: Path,
    ui_package: Path,
    fx_package: Path,
    danger_package: Path,
    g01_package: Path,
    repo: Path,
    records: list[dict[str, Any]],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    def read_formal_csv(
        source_package: Path, prefix: str, count: int
    ) -> tuple[list[dict[str, str]], str, bytes]:
        with zipfile.ZipFile(source_package) as archive:
            for entry in archive.namelist():
                if not entry.lower().endswith(".csv"):
                    continue
                data = archive.read(entry)
                try:
                    rows = list(csv.DictReader(io.StringIO(data.decode("utf-8-sig"))))
                except UnicodeDecodeError:
                    continue
                if len(rows) == count and rows and rows[0].get("资产ID", "").startswith(prefix):
                    return rows, entry, data
        raise ValueError(f"{source_package.name}: no formal {prefix} catalog with {count} rows")

    master_data, master_entry = xlsx_master_from_character_package(character_package)
    master_sha = sha256_bytes(master_data)
    workbook = openpyxl.load_workbook(io.BytesIO(master_data), data_only=False)
    tables = {
        "characters": sheet_rows(workbook, "人物清单", "CHAR"),
        "scenes": sheet_rows(workbook, "场景清单", "SCN"),
        "props": sheet_rows(workbook, "装备道具", "PROP"),
        "mech": sheet_rows(workbook, "玩法机制", "MECH"),
        "ui": sheet_rows(workbook, "UI清单", "UI"),
    }
    expected = {"characters": 71, "scenes": 91, "props": 46, "mech": 47, "ui": 83}
    observed = {key: len(value) for key, value in tables.items()}
    if observed != expected:
        raise ValueError(f"master catalog counts mismatch: {observed} != {expected}")

    with zipfile.ZipFile(character_package) as archive:
        three_view_entries = {
            chapter: next(
                name
                for name in archive.namelist()
                if PurePosixPath(name).name.startswith(f"{chapter}_")
                and "三视图" in PurePosixPath(name).name
            )
            for chapter in [f"G{number:02d}" for number in range(2, 14)]
        }
        three_view_hashes = {
            chapter: sha256_bytes(archive.read(source_entry))
            for chapter, source_entry in three_view_entries.items()
        }

    characters: list[dict[str, Any]] = []
    for index, row in enumerate(tables["characters"], start=1):
        chapter = row["星球编号"]
        name = row["人物名称"]
        characters.append(
            {
                "catalog_id": f"CAT-CHAR-{index:03d}",
                "official_id": "EDU-0077" if name == "七码" else None,
                "source_asset_id": row["资产ID"],
                "character_name": name,
                "chapter": chapter,
                "role": row.get("类别"),
                "administrator": row.get("管理员") == "是",
                "design_master_status": "complete",
                "three_view_status": "complete",
                "runtime_portrait_status": "not_produced",
                "runtime_scene_asset_status": "not_produced",
                "source_package": "PKG-CHARACTERS-V2.1",
                "source_entry": three_view_entries[chapter],
                "source_sha256": three_view_hashes[chapter],
                "registry_source_entry": master_entry,
                "registry_source_sha256": master_sha,
                "registry_status_at_V1.0": row.get("当前状态"),
            }
        )

    assets: list[dict[str, Any]] = [
        {
            "catalog_id": row["catalog_id"],
            "official_id": row["official_id"],
            "domain": "character",
            "name": row["character_name"],
            "chapter": row["chapter"],
            "scope": row["chapter"],
            "type": row["role"],
            "maturity": "design_and_three_view_complete",
            "source_package": row["source_package"],
            "source_entry": row["source_entry"],
            "source_sha256": row["source_sha256"],
            "source_granularity": "design_board_entry",
            "formal_row_id": row["source_asset_id"],
            "formal_row_entry": row["registry_source_entry"],
            "formal_row_sha256": row["registry_source_sha256"],
            "source_status": "71/71_design_identity_and_three_view_verified",
            "design_master": True,
            "production_spec": True,
            "runtime_asset": False,
            "acceptance_asset": False,
        }
        for row in characters
    ]
    formal_domains = (
        (
            "scenes",
            "scene",
            scene_package,
            "PKG-SCENES-V1.0",
            "SCN",
            91,
            normalize_scene_formal_row,
            {"name": "场景名称", "chapter": "星球编号", "scope": "星球/范围", "maturity": "状态", "priority": "优先级", "delivery_status": "交付状态"},
        ),
        (
            "props",
            "prop",
            prop_package,
            "PKG-PROPS-V3.0",
            "PROP",
            46,
            normalize_prop_formal_row,
            {"name": "道具名称", "chapter": "星球编号", "scope": "星球", "design_board": "对应正式板", "maturity": "状态", "verification_result": "核查结论"},
        ),
        (
            "mech",
            "mechanism",
            mech_package,
            "PKG-MECH-V2.0",
            "MECH",
            47,
            normalize_mech_formal_row,
            {"name": "机制名称", "mechanism_name": "机制名称", "chapter": "星球[Gxx前缀]", "scope": "星球", "maturity": "状态", "freeze_version": "冻结版本"},
        ),
        (
            "ui",
            "ui",
            ui_package,
            "PKG-UI-V2.0",
            "UI",
            83,
            normalize_ui_formal_row,
            {"name": "界面名称", "chapter": "范围", "scope": "星球", "type": "类别", "category": "类别", "batch": "分批", "independent_board_status": "独立设计板", "confirmation_status": "确认状态", "maturity": "确认状态", "freeze_version": "冻结版本"},
        ),
    )
    field_authority_map: dict[str, Any] = {
        "character": {
            "source_package": "PKG-CHARACTERS-V2.1",
            "source_entry": master_entry,
            "worksheet": "人物清单",
            "fields": {
                "formal_row_id": "资产ID",
                "name": "人物名称",
                "chapter": "星球编号",
                "type": "类别",
                "registry_status_at_V1.0": "当前状态",
            },
            "design_master_fields": "同包内十二星球人物三视图",
            "registry_reference_role": "character_identity_source_only",
        }
    }
    registry_diff_rows: list[dict[str, Any]] = []
    formal_input_records: list[dict[str, Any]] = [
        {
            "domain": "character",
            "source_package": "PKG-CHARACTERS-V2.1",
            "source_entry": master_entry,
            "source_entry_sha256": master_sha,
            "source_granularity": "registry_workbook",
        }
    ]
    for key, domain, package, package_id, prefix, count, normalizer, field_map in formal_domains:
        formal_rows, formal_entry, formal_data = read_formal_csv(package, prefix, count)
        master_ids = {row["资产ID"] for row in tables[key]}
        formal_ids = {row["资产ID"] for row in formal_rows}
        if master_ids != formal_ids:
            raise ValueError(
                f"{domain} formal catalog IDs differ from registry: "
                f"missing={sorted(master_ids - formal_ids)}, "
                f"extra={sorted(formal_ids - master_ids)}"
            )
        formal_sha = sha256_bytes(formal_data)
        formal_input_records.append(
            {
                "domain": domain,
                "source_package": package_id,
                "source_entry": formal_entry,
                "source_entry_sha256": formal_sha,
                "source_granularity": "formal_catalog_entry",
            }
        )
        field_authority_map[domain] = {
            "source_package": package_id,
            "source_entry": formal_entry,
            "fields": field_map,
            "registry_reference_role": REGISTRY_REFERENCE_ROLE,
        }
        normalized_rows = [
            normalizer(row, package_id, formal_entry, formal_sha) for row in formal_rows
        ]
        assets.extend(normalized_rows)

        registry_by_id = {row["资产ID"]: row for row in tables[key]}
        registry_name_field = {
            "scene": "场景名称",
            "prop": "名称",
            "mechanism": "机制名称",
            "ui": "名称",
        }[domain]
        formal_name_field = {
            "scene": "场景名称",
            "prop": "道具名称",
            "mechanism": "机制名称",
            "ui": "界面名称",
        }[domain]
        formal_status_field = "确认状态" if domain == "ui" else "状态"
        for formal_row in formal_rows:
            asset_id = formal_row["资产ID"]
            registry_row = registry_by_id[asset_id]
            registry_name = registry_row.get(registry_name_field)
            formal_name = formal_row.get(formal_name_field)
            registry_status = registry_row.get("当前状态")
            formal_status = formal_row.get(formal_status_field)
            differences: list[str] = []
            if registry_name != formal_name:
                differences.append("name")
            if registry_status != formal_status:
                differences.append("status")
            if differences:
                registry_diff_rows.append(
                    {
                        "asset_id": asset_id,
                        "domain": domain,
                        "registry_name": registry_name,
                        "formal_name": formal_name,
                        "registry_status": registry_status,
                        "formal_status": formal_status,
                        "difference_type": "+".join(differences),
                        "current_authority": f"{package_id}:{formal_entry}",
                        "manual_action": "否；正式清单优先，旧登记表仅保留交叉核对记录",
                    }
                )

    fx_rows, fx_entry, fx_data = read_formal_csv(fx_package, "FX-", 41)
    fx_sha = sha256_bytes(fx_data)
    formal_input_records.append(
        {
            "domain": "fx",
            "source_package": "PKG-FX-V2.0",
            "source_entry": fx_entry,
            "source_entry_sha256": fx_sha,
            "source_granularity": "formal_catalog_entry",
        }
    )
    with zipfile.ZipFile(fx_package) as archive:
        fx_files = archive.namelist()
    field_authority_map["fx"] = {
        "source_package": "PKG-FX-V2.0",
        "source_entry": fx_entry,
        "fields": {
            "name": "角色/装备",
            "chapter": "范围[Gxx前缀；无前缀为GLOBAL]",
            "scope": "范围",
            "hopa_interaction": "HOPA交互效果",
            "maturity": "状态",
            "freeze_version": "冻结版本",
        },
        "registry_reference_role": REGISTRY_REFERENCE_ROLE,
    }
    for row in fx_rows:
        require_formal_fields(
            row,
            package_id="PKG-FX-V2.0",
            source_entry=fx_entry,
            required=("资产ID", "范围", "角色/装备", "HOPA交互效果", "状态", "冻结版本"),
        )
        asset_id = row["资产ID"]
        scope = row["范围"]
        chapter = re.search(r"G\d{2}", scope or "")
        board = next(
            (
                name
                for name in fx_files
                if PurePosixPath(name).name.startswith(f"{asset_id}_")
                and name.lower().endswith(".png")
            ),
            None,
        )
        asset_number = int(asset_id[-3:])
        if board is None and 7 <= asset_number <= 16:
            board = next(
                name
                for name in fx_files
                if PurePosixPath(name).name == "FX-B装备效果设计_总览.png"
            )
        assets.append(
            {
                "catalog_id": asset_id,
                "official_id": asset_id,
                "domain": "fx",
                "name": row.get("角色/装备"),
                "chapter": chapter.group(0) if chapter else "GLOBAL",
                "scope": scope,
                "hopa_interaction": row["HOPA交互效果"],
                "maturity": row["状态"],
                "freeze_version": row["冻结版本"],
                "design_board": board,
                **formal_source_fields(
                    package_id="PKG-FX-V2.0",
                    source_entry=fx_entry,
                    source_sha256=fx_sha,
                    row_id=asset_id,
                ),
            }
        )

    danger_rows, danger_entry, danger_data = read_formal_csv(
        danger_package, "DANGER-", 76
    )
    danger_sha = sha256_bytes(danger_data)
    formal_input_records.append(
        {
            "domain": "danger",
            "source_package": "PKG-DANGER-V2.0",
            "source_entry": danger_entry,
            "source_entry_sha256": danger_sha,
            "source_granularity": "formal_catalog_entry",
        }
    )
    with zipfile.ZipFile(danger_package) as archive:
        danger_files = archive.namelist()
    field_authority_map["danger"] = {
        "source_package": "PKG-DANGER-V2.0",
        "source_entry": danger_entry,
        "fields": {
            "name": "危险名称",
            "chapter": "星球",
            "scope": "星球",
            "type": "类型",
            "application_scene": "应用场景",
            "batch": "分批",
            "independent_board_status": "独立设计板",
            "confirmation_status": "确认状态",
            "maturity": "确认状态",
            "freeze_version": "冻结版本",
        },
        "registry_reference_role": REGISTRY_REFERENCE_ROLE,
    }
    for row in danger_rows:
        require_formal_fields(
            row,
            package_id="PKG-DANGER-V2.0",
            source_entry=danger_entry,
            required=("资产ID", "星球", "危险名称", "类型", "应用场景", "分批", "独立设计板", "确认状态", "冻结版本"),
        )
        asset_id = row["资产ID"]
        scope = row["星球"]
        board = next(
            (
                name
                for name in danger_files
                if PurePosixPath(name).name.startswith(f"{asset_id}_")
                and name.lower().endswith(".png")
            ),
            None,
        )
        assets.append(
            {
                "catalog_id": asset_id,
                "official_id": asset_id,
                "domain": "danger",
                "name": row["危险名称"],
                "chapter": row["星球"],
                "scope": scope,
                "type": row["类型"],
                "application_scene": row["应用场景"],
                "batch": row["分批"],
                "independent_board_status": row["独立设计板"],
                "confirmation_status": row["确认状态"],
                "maturity": row["确认状态"],
                "freeze_version": row["冻结版本"],
                "design_board": board,
                **formal_source_fields(
                    package_id="PKG-DANGER-V2.0",
                    source_entry=danger_entry,
                    source_sha256=danger_sha,
                    row_id=asset_id,
                ),
            }
        )

    with zipfile.ZipFile(g01_package) as archive:
        g01_entry = find_entry(
            archive,
            lambda name: name.lower().endswith(".xlsx")
            and "G01序章美术资产清单_V3.0" in name,
        )
        g01_data = archive.read(g01_entry)
        g01_files = archive.namelist()
    g01_workbook = openpyxl.load_workbook(io.BytesIO(g01_data), data_only=False)
    formal_input_records.append(
        {
            "domain": "g01_addition",
            "source_package": "PKG-G01-V3.0",
            "source_entry": g01_entry,
            "source_entry_sha256": sha256_bytes(g01_data),
            "source_granularity": "formal_catalog_entry",
        }
    )
    sheet = g01_workbook["G01美术资产清单"]
    headers = [cell.value for cell in sheet[3]]
    category_board = {
        "场景": "G01序章场景概念设计总览.png",
        "道具": "G01序章道具与效果设计总览.png",
        "效果": "G01序章道具与效果设计总览.png",
        "危险视觉": "G01教学机制与危险视觉总览.png",
        "教学机制视觉": "G01教学机制与危险视觉总览.png",
    }
    field_authority_map["g01_addition"] = {
        "source_package": "PKG-G01-V3.0",
        "source_entry": g01_entry,
        "worksheet": "G01美术资产清单",
        "fields": {
            "name": "名称",
            "chapter": "constant:G01",
            "scope": "constant:G01序章",
            "type": "类别",
            "delivery_form": "交付形态",
            "freeze_requirement": "冻结要求",
            "maturity": "状态",
        },
        "registry_reference_role": "not_applicable",
    }
    g01_assets: list[dict[str, Any]] = []
    for values in sheet.iter_rows(min_row=4, values_only=True):
        if not isinstance(values[0], str):
            continue
        row = {
            str(headers[index]): value
            for index, value in enumerate(values)
            if index < len(headers) and headers[index] and value is not None
        }
        require_formal_fields(
            row,
            package_id="PKG-G01-V3.0",
            source_entry=g01_entry,
            required=("资产ID", "类别", "名称", "交付形态", "冻结要求", "状态"),
        )
        board_name = category_board[row["类别"]]
        board = next(name for name in g01_files if name.endswith(board_name))
        g01_assets.append(
            {
                "catalog_id": row["资产ID"],
                "official_id": row["资产ID"],
                "domain": "g01_addition",
                "name": row["名称"],
                "chapter": "G01",
                "scope": "G01序章",
                "type": row["类别"],
                "delivery_form": row.get("交付形态"),
                "freeze_requirement": row.get("冻结要求"),
                "maturity": row.get("状态"),
                "source_package": "PKG-G01-V3.0",
                "source_entry": g01_entry,
                "source_sha256": sha256_bytes(g01_data),
                "source_granularity": "formal_catalog_entry",
                "formal_row_id": row["资产ID"],
                "design_board": board,
                "source_status": "formal_V3.0_catalog",
                "design_master": True,
                "production_spec": True,
                "runtime_asset": False,
                "acceptance_asset": False,
            }
        )
    if len(g01_assets) != 33:
        raise ValueError(f"G01 formal asset catalog must have 33 rows, got {len(g01_assets)}")
    assets.extend(g01_assets)

    domain_counts: dict[str, int] = {}
    for row in assets:
        row["id"] = row["catalog_id"]
        domain_counts[row["domain"]] = domain_counts.get(row["domain"], 0) + 1
    expected_domains = {
        "character": 71,
        "scene": 91,
        "prop": 46,
        "fx": 41,
        "mechanism": 47,
        "danger": 76,
        "ui": 83,
        "g01_addition": 33,
    }
    if len(assets) != 488 or domain_counts != expected_domains:
        raise ValueError(f"asset catalog mismatch: total={len(assets)}, domains={domain_counts}")

    catalogs_dir = repo / "data/source/catalogs"
    report_path = repo / "docs/review/ASSET_REGISTRY_VS_FORMAL_CATALOG_DIFF.md"
    report_path.parent.mkdir(parents=True, exist_ok=True)
    report_lines = [
        "# 旧项目总登记表与正式域清单差异报告",
        "",
        f"- 生成器：`{CATALOG_GENERATOR}`",
        f"- 映射版本：`{CATALOG_MAPPING_VERSION}`",
        f"- 差异条目：{len(registry_diff_rows)}",
        "- 当前字段权威：各域正式 CSV/XLSX 的唯一正式行",
        "- 旧项目总登记表角色：`cross_check_only`，不得覆盖正式名称、状态、范围、类别或版本字段",
        "- 人工处理：正式包版本优先级明确的差异无需阻断；无法判定权威来源时导入器直接失败",
        "",
        "| asset_id | domain | registry_name | formal_name | registry_status | formal_status | difference_type | current_authority | manual_action |",
        "|---|---|---|---|---|---|---|---|---|",
    ]
    for difference in registry_diff_rows:
        values = [
            difference["asset_id"],
            difference["domain"],
            difference["registry_name"],
            difference["formal_name"],
            difference["registry_status"],
            difference["formal_status"],
            difference["difference_type"],
            difference["current_authority"],
            difference["manual_action"],
        ]
        report_lines.append(
            "| " + " | ".join(str(value or "").replace("|", "\\|").replace("\n", " ") for value in values) + " |"
        )
    report_path.write_text("\n".join(report_lines) + "\n", encoding="utf-8")
    write_json(catalogs_dir / "characters-71.json", characters)
    write_csv(catalogs_dir / "characters-71.csv", characters)
    write_json(
        catalogs_dir / "asset-catalog-488.json",
        {
            "schema_version": 3,
            "total": len(assets),
            "domain_counts": domain_counts,
            "generated_by": CATALOG_GENERATOR,
            "mapping_version": CATALOG_MAPPING_VERSION,
            "registry_reference_role": REGISTRY_REFERENCE_ROLE,
            "field_authority_map": field_authority_map,
            "source_packages": [
                "PKG-CHARACTERS-V2.1",
                "PKG-SCENES-V1.0",
                "PKG-PROPS-V3.0",
                "PKG-MECH-V2.0",
                "PKG-UI-V2.0",
                "PKG-FX-V2.0",
                "PKG-DANGER-V2.0",
                "PKG-G01-V3.0",
            ],
            "caveats": ["设计/生产母版与运行时、验收资产状态分栏记录。"],
            "assets": assets,
        },
    )
    write_csv(catalogs_dir / "asset-catalog-488.csv", assets)
    write_json(
        catalogs_dir / "master-workbook-counts.json",
        {
            "provenance_type": "multi_source_derived_catalog",
            "generated_by": CATALOG_GENERATOR,
            "mapping_version": CATALOG_MAPPING_VERSION,
            "registry_reference_role": REGISTRY_REFERENCE_ROLE,
            "field_authority_map": field_authority_map,
            "derived_from": formal_input_records,
            "counts": {**observed, "fx": 41, "danger": 76, "g01_addition": 33},
            "normalized_asset_total": 488,
            "formal_row_unique_count": len(assets),
            "registry_formal_difference_count": len(registry_diff_rows),
            "registry_formal_difference_report": repository_path(report_path),
            "design_or_production_master": True,
            "runtime_asset": False,
        },
    )
    for output in (
        catalogs_dir / "characters-71.json",
        catalogs_dir / "characters-71.csv",
    ):
        record_extraction(
            records,
            output=output,
            package_id="PKG-CHARACTERS-V2.1",
            entry=master_entry,
            source_data=master_data,
            extraction="formal_catalog_normalization",
        )
    for output in (
        catalogs_dir / "asset-catalog-488.json",
        catalogs_dir / "asset-catalog-488.csv",
        catalogs_dir / "master-workbook-counts.json",
    ):
        record_derived_catalog(
            records,
            output=output,
            derived_from=formal_input_records,
            field_authority_map=field_authority_map,
        )
    return characters, assets


def extract_art_catalog(
    package: Path,
    package_id: str,
    entry_suffix: str,
    output: Path,
    records: list[dict[str, Any]],
) -> None:
    with zipfile.ZipFile(package) as archive:
        entry = find_entry(archive, lambda name: name.endswith(entry_suffix))
        data = archive.read(entry)
    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_bytes(data)
    record_extraction(
        records,
        output=output,
        package_id=package_id,
        entry=entry,
        source_data=data,
        extraction="verbatim_csv",
    )


def extract_hopa_and_fx(
    package: Path,
    repo: Path,
    records: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    output_dir = repo / "docs/baseline/source_text/hopa"
    stats: list[dict[str, Any]] = []
    with zipfile.ZipFile(package) as archive:
        names = archive.namelist()
    docs = {
        "星骸拾荒者_HOPA玩法与技术架构冻结修正版_V1.0.docx": (
            output_dir / "HOPA_ARCHITECTURE_V1.0.md",
            "HOPA玩法与技术架构冻结修正版完整文本",
        ),
        "星骸拾荒者_FX-001星宇瞬移_HOPA交互效果详细设计_V1.0.docx": (
            output_dir / "FX-001_V1.0.md",
            "FX-001 HOPA交互效果详细设计完整文本",
        ),
    }
    for filename, (output, purpose) in docs.items():
        entry = next(name for name in names if PurePosixPath(name).name == filename)
        stats.append(
            extract_docx_entry(
                package,
                "PKG-HOPA-FX001-V1.0",
                entry,
                output,
                "V1.0",
                purpose,
                records,
            )
        )
    masters_dir = repo / "docs/baseline/production-masters/hopa-fx001"
    for filename in (
        "HOPA核心循环_冻结版.png",
        "HOPA场景分层与热点架构.png",
        "HOPA技术模块架构.png",
        "FX-001星宇瞬移_HOPA交互流程.png",
    ):
        entry = next(name for name in names if PurePosixPath(name).name == filename)
        with zipfile.ZipFile(package) as archive:
            data = archive.read(entry)
        output = masters_dir / filename
        output.parent.mkdir(parents=True, exist_ok=True)
        output.write_bytes(data)
        record_extraction(
            records,
            output=output,
            package_id="PKG-HOPA-FX001-V1.0",
            entry=entry,
            source_data=data,
            extraction="verbatim_design_production_master",
        )
    (masters_dir / "README.md").write_text(
        "# HOPA 与 FX-001 设计/生产母版\n\n"
        "本目录四张 PNG 均从已校验的 HOPA 架构与 FX-001 确认包逐字节提取，"
        "用途是设计、生产和验收参考；它们不是游戏运行时资产。\n",
        encoding="utf-8",
    )
    return stats


def extract_g01_supplemental_docs(
    package: Path,
    repo: Path,
    records: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    output_dir = repo / "docs/baseline/source_text/g01"
    with zipfile.ZipFile(package) as archive:
        entries = [
            name
            for name in archive.namelist()
            if name.lower().endswith(".docx")
            and "G01拾光号坠落之前_剧情与HOPA实施脚本" not in name
        ]
    stats: list[dict[str, Any]] = []
    for entry in entries:
        filename = PurePosixPath(entry).name
        if "G02开场边界修订_V2.2" in filename:
            output = output_dir / "G02_OPENING_BOUNDARY_V2.2.md"
            purpose = "G02开场边界修订V2.2；来源为G01 V3.0正式包内部"
        else:
            output = output_dir / f"{PurePosixPath(filename).stem}.md"
            purpose = "G01 V3.0正式包补充规范全文"
        stats.append(
            extract_docx_entry(
                package,
                "PKG-G01-V3.0",
                entry,
                output,
                "V3.0" if "V2.2" not in filename else "V2.2",
                purpose,
                records,
            )
        )
    if not any("G02开场边界修订_V2.2" in item["source_entry"] for item in stats):
        raise ValueError("G02 opening boundary V2.2 was not extracted from G01 V3.0")
    return stats


def write_completeness_report(
    repo: Path, story_stats: list[dict[str, Any]]
) -> None:
    rows = [
        "| 章节 | DOCX文本字符 | 捕获字符 | 覆盖率 | Markdown |",
        "| --- | ---: | ---: | ---: | --- |",
    ]
    for item in story_stats:
        if not re.fullmatch(r"G\d{2}", item["chapter"]):
            continue
        rows.append(
            f"| {item['chapter']} | {item['source_text_characters']} | "
            f"{item['captured_text_characters']} | "
            f"{item['coverage_ratio'] * 100:.0f}% | "
            f"`{PurePosixPath(item['output_path']).name}` |"
        )
    output = repo / "docs/story/G01-G13/COMPLETENESS_REPORT.md"
    output.write_text(
        "# G01—G13 Markdown / DOCX 完整性报告\n\n"
        "覆盖率按 DOCX `word/document.xml` 中全部 `w:t` 文本节点字符数与提取器"
        "捕获字符数核对。表格和正文均进入 Markdown；不把摘要当作全文。\n\n"
        + "\n".join(rows)
        + "\n",
        encoding="utf-8",
    )


def write_data_index(repo: Path, present: list[dict[str, Any]]) -> None:
    by_chapter = {item["chapter"]: item for item in present}
    required = [f"G{number:02d}" for number in range(1, 14)]
    missing = [chapter for chapter in required if chapter not in by_chapter]
    if missing:
        raise ValueError(f"formal chapter data missing after P0-A import: {missing}")
    index = [by_chapter[chapter] for chapter in required]
    write_json(repo / "data/source/index.json", index)


def main() -> int:
    global REPO_ROOT
    parser = argparse.ArgumentParser()
    parser.add_argument("--repo", type=Path, default=Path.cwd())
    parser.add_argument("--downloads", type=Path, required=True)
    args = parser.parse_args()
    repo = args.repo.resolve()
    REPO_ROOT = repo
    downloads = args.downloads.resolve()
    originals = repo / "source_packages/originals-or-release-links"
    paths, package_manifest = verify_and_copy_packages(downloads, originals)
    records: list[dict[str, Any]] = []
    doc_stats: list[dict[str, Any]] = []

    doc_stats.extend(
        extract_named_docs(
            paths["PKG-PRODUCT-PLAN-V1.1"],
            "PKG-PRODUCT-PLAN-V1.1",
            PRODUCT_DOCS,
            repo / "docs/baseline/source_text/product-plan",
            "confirmed versions",
            "产品、教育、计划与门禁正式全文",
            records,
        )
    )
    doc_stats.extend(
        extract_named_docs(
            paths["PKG-G02-SCRIPT-FREEZE-V1.0"],
            "PKG-G02-SCRIPT-FREEZE-V1.0",
            G02_FREEZE_DOCS,
            repo / "docs/baseline/source_text/g02-freeze",
            "confirmed versions",
            "G02阶段计划与G-SCR-01—05冻结全文",
            records,
        )
    )
    doc_stats.extend(
        extract_hopa_and_fx(paths["PKG-HOPA-FX001-V1.0"], repo, records)
    )
    doc_stats.extend(
        extract_g01_supplemental_docs(paths["PKG-G01-V3.0"], repo, records)
    )
    for source_id, output, purpose in (
        (
            "DOC-G-S2-D01-V1.0",
            repo / "docs/baseline/source_text/visual/G-S2-D01_V1.0.md",
            "S2视觉总方向关键决策冻结记录",
        ),
        (
            "DOC-G-S2-CHG-01-V1.0",
            repo / "docs/baseline/source_text/visual/G-S2-CHG-01_V1.0.md",
            "视觉路线与角色一致性修正记录",
        ),
        (
            "DOC-G-CHAR-01-V1.0",
            repo / "docs/baseline/source_text/characters/G-CHAR-01_V1.0.md",
            "主角星宇造型冻结基线",
        ),
        (
            "DOC-G-ANIM-01-V1.0",
            repo / "docs/baseline/source_text/animation/G-ANIM-01_V1.0.md",
            "图片解密游戏轻动效与骨骼动画方案",
        ),
    ):
        doc_stats.append(
            extract_standalone_docx(
                paths[source_id],
                source_id,
                output,
                purpose,
                records,
            )
        )
    story_stats, _ = extract_story_docs(
        paths["PKG-G02-G13-HOPA-V2.0"],
        paths["PKG-G01-V3.0"],
        repo,
        records,
    )
    doc_stats.extend(story_stats)

    data_status = [
        extract_complete_chapter_data(
            paths["PKG-G01-V3.0"],
            "PKG-G01-V3.0",
            "G01",
            repo,
            records,
            version="V3.0",
            g01=True,
        )
    ]
    for number in range(2, 14):
        chapter = f"G{number:02d}"
        supplemental_id = (
            f"PKG-{chapter}-DATA-V2.1" if number in (2, 3) else None
        )
        data_status.append(
            extract_complete_chapter_data(
                paths["PKG-G02-G13-DATA-V2.1"],
                "PKG-G02-G13-DATA-V2.1",
                chapter,
                repo,
                records,
                version="V2.1",
                structured_package=(
                    paths[supplemental_id] if supplemental_id else None
                ),
                structured_package_id=supplemental_id,
            )
        )
    write_data_index(repo, data_status)
    build_catalogs(
        paths["PKG-CHARACTERS-V2.1"],
        paths["PKG-SCENES-V1.0"],
        paths["PKG-PROPS-V3.0"],
        paths["PKG-MECH-V2.0"],
        paths["PKG-UI-V2.0"],
        paths["PKG-FX-V2.0"],
        paths["PKG-DANGER-V2.0"],
        paths["PKG-G01-V3.0"],
        repo,
        records,
    )
    raw_catalog_dir = repo / "data/source/catalogs/raw"
    extract_art_catalog(
        paths["PKG-SCENES-V1.0"],
        "PKG-SCENES-V1.0",
        "91场景完成清单.csv",
        raw_catalog_dir / "SCENE-91_V1.0.csv",
        records,
    )
    extract_art_catalog(
        paths["PKG-PROPS-V3.0"],
        "PKG-PROPS-V3.0",
        "46件道具正式清单.csv",
        raw_catalog_dir / "PROP-46_V3.0.csv",
        records,
    )
    extract_art_catalog(
        paths["PKG-MECH-V2.0"],
        "PKG-MECH-V2.0",
        "47项机制可视化正式清单.csv",
        raw_catalog_dir / "MECH-47_V2.0.csv",
        records,
    )
    extract_art_catalog(
        paths["PKG-UI-V2.0"],
        "PKG-UI-V2.0",
        "83项界面美术正式清单.csv",
        raw_catalog_dir / "UI-83_V2.0.csv",
        records,
    )

    write_completeness_report(repo, story_stats)
    write_json(
        repo / "source_packages/manifests/source-packages.json",
        {
            "schema_version": 1,
            "issue": 6,
            "imported": package_manifest,
            "missing_required": list(MISSING_REQUIRED),
        },
    )
    write_json(
        repo / "source_packages/manifests/missing-sources.json",
        {
            "schema_version": 1,
            "issue": 6,
            "missing_count": len(MISSING_REQUIRED),
            "items": list(MISSING_REQUIRED),
        },
    )
    write_json(
        repo / "source_packages/manifests/extracted-files.json",
        {
            "schema_version": 1,
            "count": len(records),
            "files": sorted(records, key=lambda row: row["output_path"]),
        },
    )
    write_json(
        repo / "source_packages/manifests/docx-extraction-stats.json",
        {
            "schema_version": 1,
            "documents": doc_stats,
            "all_available_documents_full_text_captured": all(
                item["coverage_ratio"] == 1 for item in doc_stats
            ),
        },
    )
    substitution_path = repo / "source_packages/manifests/substitution-map.json"
    substitution = json.loads(substitution_path.read_text(encoding="utf-8"))
    imported_rule_text = {
        "技能与装备效果HOPA正式包 V2.0": (
            "FX-001—041以已校验V2.0正式清单、章节范围和包内设计板路径为准；"
            "V1局部包仅保留为被替代资料。"
        ),
        "危险视觉HOPA正式包 V2.0": (
            "DANGER-001—076以已校验V2.0正式清单、章节范围和包内设计板路径为准；"
            "删除旧推算展开规则。"
        ),
        "G01整合正式包 V3.0": (
            "G01全文、结构化数据、33项正式资产和包内G02开场边界V2.2"
            "均以已校验V3.0原包为准。"
        ),
        "G-S2-D01 V1.0": (
            "已校验并全文提取冻结V1.0；V0.9待确认版只保留为被替代资料。"
        ),
    }
    for rule in substitution["rules"]:
        if rule["current"] in imported_rule_text:
            rule["status"] = "current_source_imported"
            rule["rule"] = imported_rule_text[rule["current"]]
    write_json(substitution_path, substitution)
    boundary = next(
        item
        for item in doc_stats
        if "G02开场边界修订_V2.2" in item["source_entry"]
    )
    (repo / "source_packages/manifests/SOURCE_IMPORT_REPORT.md").write_text(
        "# Issue #6 正式源包导入报告\n\n"
        f"- 正式源文件：{len(package_manifest)}\n"
        "- `missing_required`: 0\n"
        "- `missing_count`: 0\n"
        "- G01—G13剧情脚本：13/13正式可查询\n"
        "- G01—G13结构化数据：13/13正式可查询\n"
        "- 人物身份与三视图：71/71；运行时头像、场景资产另列为未制作\n"
        "- 资产目录：488/488\n"
        "- FX正式清单：41/41\n"
        "- DANGER正式清单：76/76\n"
        f"- G02开场边界V2.2源路径：`{boundary['source_entry']}`\n"
        f"- G02开场边界V2.2 SHA-256：`{boundary['source_sha256']}`\n"
        f"- G02开场边界V2.2 Markdown：`{boundary['output_path']}`\n"
        "- 正式运行时：HTML5/PWA + Vite + TypeScript；Unity仅作为数据来源命名保留。\n",
        encoding="utf-8",
    )
    sums = "\n".join(
        f"{item['observed_sha256']}  originals-or-release-links/{item['filename']}"
        for item in package_manifest
    )
    (repo / "source_packages/manifests/sha256sums.txt").write_text(
        sums + "\n", encoding="utf-8"
    )
    print(
        json.dumps(
            {
                "imported_packages": len(package_manifest),
                "missing_required": len(MISSING_REQUIRED),
                "extracted_files": len(records),
                "documents": len(doc_stats),
            },
            ensure_ascii=False,
        )
    )
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except Exception as error:
        print(f"source import failed: {error}", file=sys.stderr)
        raise
